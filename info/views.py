import requests
from bs4 import BeautifulSoup
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
import urllib3
from openpyxl.utils import get_column_letter
import xlsxwriter 
import re
from datetime import datetime
import locale
import json
import sys

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

sys.stdout.reconfigure(encoding='utf-8')

try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    pass

def datos_valparaiso(url):
    html_texto = requests.get(url, verify=False).text
    soup = BeautifulSoup(html_texto, 'html.parser')
    
    sitios = []
    #for i in range(7, 15): 
    for i in range(7, 10): 
        sitio_div = soup.find("div", class_=f"pln-titulo{i}")
        if sitio_div:
            sitio_nombre = sitio_div.find("span").text.strip() 
            sitios.append(sitio_nombre)
    
    datos = []
    
    fecha = []
    for i in range(7):  
        cellinfo = soup.find("div", class_=f"cellinfo-{i}-0")  
        fecha_result = "fecha no disponible"
        
        if cellinfo:
            dia_element = cellinfo.find("span", class_="text-dark pln-cell-fecha")
            mes_element = dia_element.find_next("span", class_="text-dark pln-cell-fecha") if dia_element else None
            
            if dia_element and mes_element:
                fecha_result = f"{dia_element.text.strip()} {mes_element.text.strip()}"
        
        fecha.append(fecha_result)

    for fila_idx in range(7): 
        #for columna_idx in range(0, 9):
        for columna_idx in range(0, 4):  
            cellinfo = soup.find("div", class_=f"cellinfo-{fila_idx}-{columna_idx}")
            
            nombre_nave = ""
            hora = ""
            posicion = ""
            
            if cellinfo:
                nombre_nave_element = cellinfo.find("span", class_="pln-nombre-nave")
                posicion_element = cellinfo.find("span", class_="pln-posicion")
                hora_element = cellinfo.find("span", class_="pln-cell-hora text-primary")
                
                nombre_nave = nombre_nave_element.text.strip() if nombre_nave_element else "N/A"
                posicion = posicion_element.text.strip() if posicion_element else "N/A"
                hora = hora_element.text.strip() if hora_element else "N/A"

            datos.append({
                "Nombre Nave": nombre_nave,
                "Fecha": fecha[fila_idx],  
                "Hora": hora,
                "Posición": posicion,
                "Sitio": sitios[columna_idx - 1] if columna_idx - 1 < len(sitios) else "Sin Sitio"
            })
            
    return [nave for nave in datos if nave["Nombre Nave"] != "N/A"]

def limpiar_json(datos):
    naves_menor_fecha = {}

    for d in datos:
        if not d['nave']:  
            continue

        if d['metros']:
            d['metros'] = re.sub(r'^0+', '', d['metros']) 

        fecha_texto = d['fecha']
        if fecha_texto:
            try:
                dia = int(fecha_texto.split()[0])  
                
                fecha_comparable = datetime(2024, 2, dia) 

                if d['nave'] not in naves_menor_fecha or fecha_comparable < naves_menor_fecha[d['nave']]['fecha_comparable']:
                    naves_menor_fecha[d['nave']] = {**d, 'fecha_comparable': fecha_comparable}

            except ValueError:
                continue 

    datos_filtrados = [{k: v for k, v in nave.items() if k != 'fecha_comparable'} for nave in naves_menor_fecha.values()]

    return datos_filtrados

def datos_san_antonio(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "es-ES,es;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Referer": "https://gessup.puertosanantonio.com/Planificaciones/general.aspx",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1"
        }

        response = requests.get(url, headers=headers, verify=False, timeout=15)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        fechas = soup.find_all('td', class_=re.compile(r'titulo', re.I))
        fechas_texto = [fecha.get_text(strip=True).replace('\n', '') for fecha in fechas]

        if not fechas_texto:
            print("No se encontraron fechas con el selector CSS especificado. Revisa el HTML.")
            return []

        print(f"Fechas encontradas: {fechas_texto}")

        contenedor_planificacion = soup.find('table', class_=re.compile(r'planificacion', re.I))
        if not contenedor_planificacion:
            print("No se encontró el contenedor principal de planificación.")
            return []

        tablas = contenedor_planificacion.select('tr > td > table')
        if not tablas:
            print("No se encontraron tablas dentro de '.planificacion > tbody > tr > td > table'. Revisa el HTML.")
            return []

        datos = []
        fecha_index = 0
        celdas_por_fecha = 7

        for i, tabla in enumerate(tablas):
            filas = tabla.find_all('tr')
            for fila in filas:
                celdas = fila.find_all('td')
                # for celda in celdas:
                for j, celda in enumerate(celdas[:5]):
                    texto = celda.get_text(strip=True).replace('\n', '')
                    if texto:
                        hora = re.search(r'(\d{2}:\d{2})', texto)
                        metros = re.search(r'(\d+\.?\d*)m', texto)
                        nave = re.sub(r'(\d{2}:\d{2})|(\d+\.?\d*m)', '', texto).strip().upper()

                        if hora or metros or nave:
                            datos.append({
                                'fecha': fechas_texto[fecha_index] if fecha_index < len(fechas_texto) else None,
                                'hora': hora.group(0) if hora else None,
                                'metros': metros.group(0) if metros else None,
                                'nave': nave if nave else None
                            })
            if (i + 1) % celdas_por_fecha == 0 and fecha_index + 1 < len(fechas_texto):
                fecha_index += 1

        datos_limpios = limpiar_json(datos)
        print(json.dumps(datos_limpios, indent=4, ensure_ascii=False))
        return datos_limpios

    except SystemExit as se:
        print(f"SystemExit capturado: {se}")
        return []
    except requests.exceptions.RequestException as e:
        print(f"Error en la solicitud HTTP: {e}")
        return []
    except Exception as e:
        print(f"Ocurrió un error: {e}")
        return []

def cargar_datos(opcion):
    if opcion == "Valparaíso":
        url = "https://pln.puertovalparaiso.cl/pln/"
        try:
            response = requests.get(url, verify=False)
            response.raise_for_status() 
            datos = datos_valparaiso(url)
            return datos, "Nombre Nave"
        except requests.exceptions.RequestException as e:
            print(f"Error al acceder a la página de Valparaíso: {e}")
            return [], "Error de conexión a Valparaíso"
    
    elif opcion == "San Antonio":
        url = "https://gessup.puertosanantonio.com/Planificaciones/general.aspx"
        try:
            response = requests.get(url, verify=False)
            response.raise_for_status()
            datos = datos_san_antonio(url)
            return datos, "nave"
        except requests.exceptions.RequestException as e:
            print(f"Error al acceder a la página de San Antonio: {e}")
            return [], "Error de conexión a San Antonio"

    return [], ""

def index(request):
    if request.method == "POST":
        puerto = request.POST.get('puerto', 'Valparaíso')
    else:
        puerto = request.GET.get('puerto', 'Valparaíso')
        
    datos, clave = cargar_datos(puerto)

    if 'selected_ships' not in request.session:
        request.session['selected_ships'] = {}
    global_selected_ships = request.session['selected_ships']
    selected_ships = global_selected_ships.get(puerto, [])

    if request.method == "POST":
        try:
            selected_indices = [int(idx) for idx in request.POST.getlist('selected_ship')]
        except ValueError:
            selected_indices = []
        global_selected_ships[puerto] = selected_indices
        request.session['selected_ships'] = global_selected_ships
        selected_ships = selected_indices

    context = {
        'puerto': puerto,
        'datos': datos,
        'clave': clave,
        'selected_ships': selected_ships,
    }
    return render(request, 'info/index.html', context)

def detalle(request, index):
    puerto = request.GET.get('puerto', 'Valparaíso')
    datos, clave = cargar_datos(puerto)
    
    try:
        elemento = datos[index]
    except IndexError:
        return JsonResponse({"error": "Elemento no encontrado"}, status=404)
    
    return JsonResponse({
        'puerto': puerto,
        'elemento': elemento,
    })

def parse_fecha(fecha_str, origen="valparaiso"):
    try:
        parts = fecha_str.split()
        if origen == "valparaiso":
            if len(parts) >= 2:
                day = parts[1]
            else:
                return None
        else:  
            if len(parts) >= 1:
                day = parts[0]
            else:
                return None
        day_int = int(day)
        now = datetime.now()
        dt = datetime(year=now.year, month=now.month, day=day_int)
        return dt
    except Exception:
        return None
    
def datos_valparaiso_anunciadas(url):
    html_texto = requests.get(url, verify=False).text
    soup = BeautifulSoup(html_texto, 'html.parser')
    
    rows = soup.select('tbody tr')

    extracted_data = []
    current_nave = None

    for row in rows:
        divs = row.select('.fila-estrecha > div')
        text_content = " ".join([div.get_text(" ", strip=True) for div in divs])

        nave_match = re.findall(r'\b[A-Z\s]+\b', text_content)
        if nave_match:
            current_nave = nave_match[0].strip()

        fecha, hora, ps = "No disponible", "No disponible", "No disponible"
        
        fecha_hora_match = re.findall(r'(\d{2}/\d{2}/\d{2}) (\d{2}:\d{2})', text_content)
        if fecha_hora_match:
            fecha, hora = fecha_hora_match[0]

        ps_match = re.search(r'PS:(\d{2}/\d{2}/\d{2} \d{2}:\d{2})', text_content)
        if ps_match:
            ps = ps_match.group(1)

        if current_nave:
            extracted_data.append({
                "Nave": current_nave,
                "Fecha": fecha,
                "Hora": hora,
                "PS": ps
            })

    return extracted_data

def datos_san_antonio_anunciadas(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "es-ES,es;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Referer": "https://gessup.puertosanantonio.com/Planificaciones/general.aspx",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1"
        }

        response = requests.get(url, headers=headers, verify=False, timeout=14)
        response.raise_for_status()  

        soup = BeautifulSoup(response.text, 'html.parser')

        encabezados = soup.find('tr', class_='GridViewHeader').find_all('th')
        encabezado_texto = [encabezado.text.strip() for encabezado in encabezados]

        filas = soup.find_all('tr', class_=['GridView', 'GridViewAlternativa'])

        datos = []

        for fila in filas:
            columnas = fila.find_all('td')
            if len(columnas) >= len(encabezado_texto):  
                fila_datos = {encabezado_texto[i]: columnas[i].text.strip() for i in range(len(encabezado_texto))}
                eta = fila_datos.get("E.T.A.", "").strip()
                nave = fila_datos.get("Nave", "").strip()

                if eta and nave:
                    datos.append({
                        "E.T.A.": eta,
                        "Nave": nave
                    })

        return datos

    except requests.exceptions.RequestException as e:
        print(f"Error al realizar la solicitud HTTP: {e}")
    except Exception as e:
        print(f"Error al procesar los datos de San Antonio: {e}")
    
    return []  

def descargar_excel_naves_anunciadas(request):
    print("Generando archivo Excel para naves anunciadas...")

    url_valparaiso_anunciadas = "https://pln.puertovalparaiso.cl/pln/"
    url_san_antonio_anunciadas = "https://gessup.puertosanantonio.com/Planificaciones/general.aspx"

    datos_naves_anunciadas_valpo = datos_valparaiso_anunciadas(url_valparaiso_anunciadas)
    datos_naves_anunciadas_sa = datos_san_antonio_anunciadas(url_san_antonio_anunciadas)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=naves_anunciadas.xlsx'

    workbook = xlsxwriter.Workbook(response)

    ws_naves_anunciadas_valpo = workbook.add_worksheet("Valparaíso-NavesAnunciadas")
    encabezados_anunciadas_valpo = ["Nave", "Fecha", "Hora", "PS"]
    ws_naves_anunciadas_valpo.write_row('A1', encabezados_anunciadas_valpo)

    for i, nave in enumerate(datos_naves_anunciadas_valpo, start=1):
        ws_naves_anunciadas_valpo.write(i, 0, nave.get("Nave", "Sin información"))
        ws_naves_anunciadas_valpo.write(i, 1, nave.get("Fecha", "No disponible"))
        ws_naves_anunciadas_valpo.write(i, 2, nave.get("Hora", "No disponible"))
        ws_naves_anunciadas_valpo.write(i, 3, nave.get("PS", "No disponible"))

    ws_naves_anunciadas_sa = workbook.add_worksheet("SanAntonio-NavesAnunciadas")
    encabezados_anunciadas_sa = ["E.T.A.", "Nave"]
    ws_naves_anunciadas_sa.write_row('A1', encabezados_anunciadas_sa)

    for i, nave in enumerate(datos_naves_anunciadas_sa, start=1):
        ws_naves_anunciadas_sa.write(i, 0, nave.get("E.T.A.", "No disponible"))
        ws_naves_anunciadas_sa.write(i, 1, nave.get("Nave", "Sin información"))

    workbook.close()
    return response

def descargar_excel(request):
    print("Entrando en la vista descargar_excel...")
    
    if 'descargar_excel' in request.POST:
        print("Formulario recibido con la opción de descarga.")

        global_selected_ships = request.session.get('selected_ships', {})
        seleccionados_valparaiso = global_selected_ships.get('Valparaíso', [])
        seleccionados_san_antonio = global_selected_ships.get('San Antonio', [])
        
        if not seleccionados_valparaiso and not seleccionados_san_antonio:
            print("No hay naves seleccionadas.")
            return HttpResponse("No hay naves seleccionadas.", status=400)
        
        datos_seleccionados_valparaiso = []
        datos_seleccionados_san_antonio = []

        datos_valparaiso, clave_valparaiso = cargar_datos("Valparaíso")
        for idx in seleccionados_valparaiso:
            if idx < len(datos_valparaiso):
                datos_seleccionados_valparaiso.append(datos_valparaiso[idx])

        datos_san_antonio, clave_san_antonio = cargar_datos("San Antonio")
        for idx in seleccionados_san_antonio:
            if idx < len(datos_san_antonio):
                datos_seleccionados_san_antonio.append(datos_san_antonio[idx])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=naves_seleccionadas.xlsx'
        
        workbook = xlsxwriter.Workbook(response)
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})

        ws_valparaiso = workbook.add_worksheet("Valparaíso")
        encabezados_valparaiso = ["Nombre Nave", "Fecha", "Hora"]
        ws_valparaiso.write_row('A1', encabezados_valparaiso)

        for i, nave in enumerate(datos_seleccionados_valparaiso, start=1):
            nombre = nave.get("Nombre Nave", "Pending")
            fecha_str = nave.get("Fecha", "Pending")
            hora = nave.get("Hora", "Pending")

            fecha_dt = parse_fecha(fecha_str, origen="valparaiso")
            
            ws_valparaiso.write(i, 0, nombre)
            if fecha_dt:
                ws_valparaiso.write_datetime(i, 1, fecha_dt, date_format)
            else:
                ws_valparaiso.write(i, 1, '')
            ws_valparaiso.write(i, 2, hora)

        ws_valparaiso.set_tab_color('green')
        
        ws_sanantonio = workbook.add_worksheet("San Antonio")
        encabezados_sanantonio = ["Nombre Nave", "Fecha", "Hora"]
        ws_sanantonio.write_row('A1', encabezados_sanantonio)

        for i, nave in enumerate(datos_seleccionados_san_antonio, start=1):
            nombre = nave.get("nave", "Pending")
            fecha_str = nave.get("fecha", "Pending")
            hora = nave.get("hora", "Pending")

            fecha_dt = parse_fecha(fecha_str, origen="san_antonio")

            ws_sanantonio.write(i, 0, nombre)
            if fecha_dt:
                ws_sanantonio.write_datetime(i, 1, fecha_dt, date_format)
            else:
                ws_sanantonio.write(i, 1, '')
            ws_sanantonio.write(i, 2, hora)

        ws_sanantonio.set_tab_color('blue')

        workbook.close()
        return response
    else:
        print("Solicitud no válida")
        return HttpResponse("Solicitud no válida", status=400)

def seleccionar_naves(request):
    if request.method == "POST":
        seleccionados_valores = request.POST.getlist("selected_ship")
        seleccionados = []
        for valor in seleccionados_valores:
            try:
                puerto, idx_str = valor.split("-", 1)
                idx = int(idx_str)
                datos, clave = cargar_datos(puerto)
                nave = datos[idx]
                nave['Puerto'] = puerto
                seleccionados.append(nave)
            except (ValueError, IndexError):
                continue  

        request.session['selected_ships'] = seleccionados

        if "descargar_excel" in request.POST:
            return descargar_excel(request)

        context = {'seleccionados': seleccionados}
        return render(request, 'info/seleccionados.html', context)
    else:
        datos_val, clave_val = cargar_datos("Valparaíso")
        datos_sa, clave_sa = cargar_datos("San Antonio")
        context = {
            'datos_val': datos_val,
            'clave_val': clave_val,
            'datos_sa': datos_sa,
            'clave_sa': clave_sa,
        }
        return render(request, 'info/seleccionar.html', context)

def eliminar_nave(request, puerto, idx):
    global_selected_ships = request.session.get('selected_ships', {})
    selected_list = global_selected_ships.get(puerto, [])
    if idx in selected_list:
        selected_list.remove(idx)
        global_selected_ships[puerto] = selected_list
        request.session['selected_ships'] = global_selected_ships
    return redirect(f"/?puerto={puerto}")

def check_updates(request):
    puerto = request.GET.get('puerto', 'Valparaíso')
    datos, clave = cargar_datos(puerto)
    
    global_selected_ships = request.session.get('selected_ships', {})
    selected_ships = global_selected_ships.get(puerto, [])

    if 'last_info' not in request.session:
        request.session['last_info'] = {}
    last_info = request.session['last_info']
    
    updates = []
    for idx in selected_ships:
        if idx < len(datos):
            current_ship = datos[idx]
            key = f"{puerto}-{idx}"
            if key in last_info and last_info[key] != current_ship:
                updates.append(current_ship)

            last_info[key] = current_ship
            
    request.session['last_info'] = last_info
    return JsonResponse({'updates': updates})